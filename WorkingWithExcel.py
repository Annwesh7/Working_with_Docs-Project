from openpyxl import Workbook, load_workbook

# Loading a workbook
workbook = load_workbook('Student_Scores.xlsx')

# Accessing current sheet in the workbook
worksheet = workbook.active                         # The active function gets the sheet which is currently being used
print(worksheet)                                    # The name of worksheet gets printed

# Accessing cells
print(worksheet['A1'])

# Accessing cell values
print(worksheet['A1'].value)

# Changing/Assigning values in the cells
worksheet['A1'].value = 'Student Name'              # Earlier the value in A1 was Name
# or another way of assigning values is
worksheet['A1'] = 'Student Name'
workbook.save('Student_Scores.xlsx')                # We need to save the workbook otherwise changes will not appear

# Printing names of all sheets
print(workbook.sheetnames)

# Accessing the desired sheet
worksheet = workbook['Scores']
print(worksheet)

# Creating a new worksheet
workbook.create_sheet('Progress report')
print(workbook.sheetnames)
workbook.save('Student_Scores.xlsx')

# Appending values in the Progress report sheet
worksheet = workbook.active
worksheet.title = 'Progress report'
worksheet.append(["Ram has a positive progress report"])
worksheet.append(["Shyam needs to improve in Python and Excel"])
worksheet.append(["Sita has an excellent progress report"])
worksheet.append(["Gita's progress report is good but needs to improve in SQL"])
worksheet.append(["Karan is doing good, just needs some improvement in Python"])
worksheet.append(["Arjun's overall report is good"])
workbook.save('Student_Scores.xlsx')

# Accessing multiple cells
worksheet = workbook.active
worksheet.title = 'Sheet2'
for row in range(1, 8):
    for column in range(1, 5):
        char = chr(65 + column)
        print(worksheet[char + str(row)])
        # To print values we can write:
        # print(worksheet[char+str(row)].value)
workbook.save('Student_Scores.xlsx')

# Inserting rows
worksheet = workbook.active
worksheet.insert_rows('5')                      # To insert a row at the specified row number
worksheet.delete_rows('5')                      # To delete a row from the specified row number

# Inserting columns
worksheet = workbook.active
worksheet.insert_cols('5')                      # To insert a column at specified column number
worksheet.delete_cols('5')                      # To delete a column from specified column number





