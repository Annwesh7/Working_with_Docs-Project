from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

Scores = { "Ram": {"SQL Score": 90, "Tableau Score": 85, "Python Score": 80, "Excel Score": 75},
           "Shyam": {"SQL Score": 85, "Tableau Score": 75, "Python Score": 65, "Excel Score": 55},
           "Sita": {"SQL Score": 95, "Tableau Score": 90, "Python Score": 85, "Excel Score": 90},
           "Gita": {"SQL Score": 65, "Tableau Score": 75, "Python Score": 80, "Excel Score": 85},
           "Karan": {"SQL Score": 85, "Tableau Score": 90, "Python Score": 65, "Excel Score": 75},
           "Arjun": {"SQL Score": 70, "Tableau Score": 80, "Python Score": 85, "Excel Score": 90}, }

# Creating a new workbook
workbook = Workbook()
worksheet = workbook.active

# Defining headings and adding them to the worksheet
headings = ["Name"] + list(Scores["Ram"].keys())
worksheet.append(headings)
worksheet.title = "Marks"

# Adding marks of candidates in the worksheet
for candidate in Scores:
    marks = list(Scores[candidate].values())
    worksheet.append([candidate] + marks)

# Styling cells
for column in range(1, 6):
    worksheet[get_column_letter(column) + "1"].font = Font(bold=True, color="00993300")

# Saving the workbook
workbook.save('Marks.xlsx')

